import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart3D, Reference

# inlezen csv
# ingeven padnaam naar csv, en indien van toepassing aanvullen delimiter
lijst = pd.read_csv(r"ttt4.csv", delimiter=';')
print(lijst)

# identificeer wanneer objectnaam 'leeg' is
ontbrekende_objectnaam = pd.isna(lijst['objectnaam'])
ontbrekende_objectnaam = lijst[ontbrekende_objectnaam]
print(ontbrekende_objectnaam)

# 1.2 Titel

ontbrekende_titel = pd.isna(lijst['titel'])
ontbrekende_titel = lijst[ontbrekende_titel]
print(ontbrekende_titel)

# 1.3 Beschrijving

ontbrekende_beschrijving = pd.isna(lijst['beschrijving'])
ontbrekende_beschrijving = lijst[ontbrekende_beschrijving]
print(ontbrekende_beschrijving)

# 1.4 Vervaardiger

ontbrekende_vervaardiger = pd.isna(lijst['vervaardiger'])
ontbrekende_vervaardiger = lijst[ontbrekende_vervaardiger]
print(ontbrekende_vervaardiger)

ontbrekende_vervaardiger_rol = pd.isna(lijst['vervaardiger.rol'])
ontbrekende_vervaardiger_rol = lijst[ontbrekende_vervaardiger_rol]
print(ontbrekende_vervaardiger_rol)

ontbrekende_vervaardiger_plaats = pd.isna(lijst['vervaardiging.plaats'])
ontbrekende_vervaardiger_plaats = lijst[ontbrekende_vervaardiger_plaats]
print(ontbrekende_vervaardiger_plaats)

# 1.5 Datering

ontbrekende_datering_begin = pd.isna(lijst['vervaardiging.datum.begin'])
ontbrekende_datering_begin = lijst[ontbrekende_datering_begin]
print(ontbrekende_datering_begin)

ontbrekende_datering_eind = pd.isna(lijst['vervaardiging.datum.eind'])
ontbrekende_datering_eind = lijst[ontbrekende_datering_eind]
print(ontbrekende_datering_eind)

wb = Workbook()

# maak sheet (tabblad)
ws = wb.create_sheet("Objectnaam")
# zet dataframe (pandas) om naar rijen in het tabblad
rows = dataframe_to_rows(ontbrekende_objectnaam, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Titel")
rows = dataframe_to_rows(ontbrekende_titel, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Beschrijving")
rows = dataframe_to_rows(ontbrekende_beschrijving, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Vervaardiger")
rows = dataframe_to_rows(ontbrekende_vervaardiger, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Vervaardigerrol")
rows = dataframe_to_rows(ontbrekende_vervaardiger_rol, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Vervaardigingplaats")
rows = dataframe_to_rows(ontbrekende_vervaardiger_plaats, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Dateringbegin")
rows = dataframe_to_rows(ontbrekende_datering_begin, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet("Dateringeind")
rows = dataframe_to_rows(ontbrekende_datering_eind, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"C:\Users\teugelso\PycharmProjects\pythonProject\TraintheTrainer_04\output.xlsx")

# 1. tellen van de data (pandas)

# tel aantal keer objectnummer aanwezig is in lijst records zonder objectnamen (gezien objectnummer altijd aanwezig)
ontbrekende_objectnamen = ontbrekende_objectnaam['objectnummer'].count()
# doe hetzelfde voor de overige velden
ontbrekende_titels = ontbrekende_titel['objectnummer'].count()
ontbrekende_beschrijvingen = ontbrekende_beschrijving['objectnummer'].count()
ontbrekende_vervaardigers = ontbrekende_vervaardiger['objectnummer'].count()
ontbrekende_dateringen_begin = ontbrekende_datering_begin['objectnummer'].count()
ontbrekende_dateringen_eind = ontbrekende_datering_eind['objectnummer'].count()
ontbrekende_vervaardigers_rol = ontbrekende_vervaardiger_rol['objectnummer'].count()
ontbrekende_vervaardigers_plaats = ontbrekende_vervaardiger_plaats['objectnummer'].count()

print(ontbrekende_vervaardigers)

labels = ["objectnaam", "titel", "beschrijving", "vervaardiger", "vervaardigerrol", "vervaardigingplaats", "dateringbegin", "dateringeind"]
ontbrekende_data = [ontbrekende_objectnamen, ontbrekende_titels, ontbrekende_beschrijvingen, ontbrekende_vervaardigers, ontbrekende_vervaardigers_rol, ontbrekende_vervaardigers_plaats,
                    ontbrekende_dateringen_begin, ontbrekende_dateringen_eind]
ws = wb.active
ws.title = 'Info'
ws['A1'] = "Ontbrekende velden"
rij1 = 2
for label in labels:
    ws.cell(row=rij1, column=1).value = label
    rij1 += 1
rij2 = 2
for veld in ontbrekende_data:
    ws.cell(row=rij2, column=2).value = veld
    rij2 += 1
# selecteer de data
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=8)

# selecteer de labels
labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=8)

# creeer grafiek
chart = BarChart3D()

# voeg grafiek toe aan excel
ws.add_chart(chart, "E2")

chart.title = 'Ontbrekende data'

# voeg namen x en y as toe
chart.y_axis.title = 'aantal'
chart.x_axis.title = 'velden'

# voeg de data & labels toe
chart.add_data(data)
chart.set_categories(labels)

wb.save(r"C:\Users\teugelso\PycharmProjects\pythonProject\TraintheTrainer_04\output.xlsx")